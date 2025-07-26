# usage: xl2csv.py [-h] [-nc] [-s [SHEET ...]] [-f [FILTER ...]] [files ...]

# Convert one or more Excel (.xlsx) workbooks into unified CSV output.

# positional arguments:
#   files                 pattern(s) to match input filepaths. Note: considers only .xlsx files.

# options:
#   -h, --help            show this help message and exit
#   -nc, --no-context     Dont Include `workbook_name` and `sheet_name` columns in every output row (default OFF).
#   -s [SHEET ...], --sheet [SHEET ...]
#                         pattern(s) to match sheet names inside workbook
#   -f [FILTER ...], --filter [FILTER ...]
#                         pattern(s) to match cell values inside worksheets and filter

import argparse
parser = argparse.ArgumentParser(description='Convert one or more Excel (.xlsx) workbooks into unified CSV output.')
parser.add_argument('files', help='pattern(s) to match input filepaths. Note: considers only .xlsx files.', default=None, nargs='*')
parser.add_argument('-nc', '--no-context', action='store_true', help='Dont Include `workbook_name` and `sheet_name` columns in every output row (default OFF).')
parser.add_argument('-s', '--sheet', type=str, default='*', nargs='*', help='pattern(s) to match sheet names inside workbook', )
parser.add_argument('-f', '--filter', type=str, default='*', nargs='*', help='pattern(s) to match row values inside worksheets and filter', )

parser.usage = parser.format_usage().replace('xl2csv.py', 'xl2csv')

# this is where help gets print.
arg_dict:dict[str, list[str]] = vars(parser.parse_args())

# optimizing
import re
if arg_dict['filter']=='*':
    pattern_matching = []
else:
    pattern_matching = [re.compile(pattern) for pattern in arg_dict['filter']]

if arg_dict['files']:
    from glob import glob 
    xlFilePaths = [file for pattern in arg_dict['files'] for file in glob(pattern) if file.endswith('.xlsx')]    
    if not xlFilePaths:
        print('No workbooks found with given path patterns')
else:
    print('No Excel files provided. Run xl2csv --help for usage instructions.')
    exit(0)

from openpyxl import load_workbook
from fnmatch import fnmatchcase

import os
import tempfile
import shutil

def runTemp(source_file_path, func, *args, **kwargs):
    with tempfile.TemporaryDirectory() as temp_dir:
        dest_file_path = os.path.join(temp_dir, os.path.basename(source_file_path))
        shutil.copy2(source_file_path, dest_file_path)
        return func(dest_file_path, *args, **kwargs)

import re 
def xl2csv(file):
    """
    Read current temporary copy of workbook
    handle empty rows as null strings -- this is by design
    """
    # arguments from shell
    global arg_dict

    # workbook = runTemp(file, load_workbook, read_only=True, data_only=True)
    workbook = load_workbook(file, read_only=True, data_only=True)
    writer = csv.writer(sys.stdout, delimiter=',', lineterminator='\n')
    
    try:
        selected_sheets = [sheet_name for sheet_name in workbook.sheetnames if any(fnmatchcase(sheet_name, sheet_pattern) for sheet_pattern in arg_dict['sheet'])]
        
        for sheet_name in selected_sheets:
            worksheet = workbook[sheet_name]

            if arg_dict['no_context']:
                row_prefix = []
            else:
                row_prefix = [os.path.basename(file) + ':' +sheet_name]
                
            # Check if worksheet has any data
            if worksheet.max_row is None or worksheet.max_column is None:
                # dont print anything ? 
                writer.writerow(row_prefix + [''])
                continue
            
            # Get all data using iter_rows
            for row in worksheet.iter_rows(values_only=True):
                # Handle completely empty rows
                record = ['' if cell is None else str(cell) for cell in row]
                if not pattern_matching:
                    writer.writerow(row_prefix + record)
                else:
                    # if any(pattern.search(field) for field in row):
                    if any(p.search(field or '') for p in pattern_matching for field in record):
                        writer.writerow(row_prefix + record)
        return 
    finally:
        workbook.close()

import csv 
import sys 
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
for xlFile in xlFilePaths:
    xl2csv(xlFile)
