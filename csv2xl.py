# usage: csv2xl [options] <input-csv-file(s)>
# Convert CSV data into one or more Excel (.xlsx) workbooks and worksheets.
# If `workbook_name` and `sheet_name` columns are present and mapped, splits data into respective files and sheets.

# Options:
#   -h, --help            Show this help message and exit.
#   -s, --split           uses first 2 rows as workbook name and sheet name and splits accordingly
#                         Note: Raises error if number of fields < 2 for any rows or first field does not end with .xlsx
#                         Default value OFF (Writes all output to 1 workbook and 1 sheet)
#   -o, --output DIR      Specify output directory.
#                         Defaults to current directory.
#                         Note: default workbook name is Book{n}.xlsx integer (n) will be set so as to not overwrite existing workbooks in given path
#   -w, --overwrite       Overwrites if needed (Does not use integer naming)
#                
# Examples:
#   # Rebuild XLSX files from CSV with workbook_name and sheet_name fields as routing directives
#   csv2xl data_with_context.csv --map-context workbook_name,sheet_name --output ./output_dir/
#   # Convert plain CSV to a single XLSX file ignoring context columns
#   csv2xl flat_data.csv --literal --output report.xlsx

import argparse
parser = argparse.ArgumentParser(description='Convert piped CSV data into one or more Excel (.xlsx) workbooks and worksheets.')
parser.add_argument('-s', '--split', action='store_true', help='if first field is given as workbook_name:sheet_name then splits data accordingly into workbooks and sheets, Rasies error if colon is missing in first field or row is empty.')
parser.add_argument('-o', '--output', type=str, default='.', help='Output dir (defaults to current directory)')
parser.add_argument('-w', '--overwrite', action='store_true', help='pattern(s) to match row values inside worksheets and filter', )

usage_str = parser.format_usage()
parser.usage = '(some commands that print to stdout) |' + usage_str[6:].replace('csv2xl.py', 'csv2xl')

arg_dict = vars(parser.parse_args())


WB_DEFAULT = 'Book.xlsx'
ST_DEFAULT = 'Sheet'

import sys
import csv 

if sys.stdin.isatty():
    # parser.
    
    parser.print_usage()
    exit(0)
else:
    pass

# contains dict[str, set] = {workbook:{sheet1, sheet2}}
meta_data:dict[str, dict] = {}

def convert(field: str):
    """Convert field to appropriate type (int, float, or string)."""
    field = field.strip()
    if not field:
        return ""
    if field.isnumeric():
        return int(field)
    else:
        try:
            res = float(field)
            return res
        except ValueError:
            return field

def createWorkBook(file, sheet):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    meta_data[file] = {}
    meta_data[file]['object'] = wb
    meta_data[file]['sheets'] = {}
    meta_data[file]['sheets'][sheet] = {}
    meta_data[file]['sheets'][sheet]['object'] = ws
    meta_data[file]['sheets'][sheet]['write_row'] = 1
    
def createSheet(file, sheet):
    wb:Workbook = meta_data[file]['object']
    ws = wb.create_sheet(title=sheet)
    meta_data[file]['sheets'][sheet] = {}
    meta_data[file]['sheets'][sheet]['object'] = ws
    meta_data[file]['sheets'][sheet]['write_row'] = 1

import os
def getSaveFilename(root_dir, filename_ext:str):
    if '.' in filename_ext:
        filename, ext = filename_ext[:filename_ext.rfind('.')], filename_ext[filename_ext.rfind('.'):]
        cut_start = len(filename)
        cut_end = -len(ext)
    else:
        filename, ext = filename_ext, ''
        cut_start = len(filename)
        cut_end = None

    # check all prexisting filenames
    from glob import glob
    existing_filenames = [f for f in glob(filename + '*' + ext, root_dir=root_dir) if f[cut_start: cut_end].isnumeric()]
    
    existing_numeric_suffixes = [int(existing_file[cut_start: cut_end]) for existing_file in existing_filenames]
    max_idx = max(existing_numeric_suffixes) if existing_numeric_suffixes else 0

    if os.path.exists(os.path.join(root_dir, filename_ext)):
        existing_filenames = [filename_ext] + existing_filenames

    if not existing_filenames:
        return existing_filenames, filename_ext

    new_filename = filename + f'{max_idx+1}' + ext
    return existing_filenames, new_filename

def save_all_workbooks():
    """Save all created workbooks to files."""
    for filename, data in meta_data.items():
        wb: Workbook = data['object']
        existing_filenames, new_filename = getSaveFilename(arg_dict['output'], filename)
        if not new_filename == filename:
            if arg_dict['overwrite'] and os.path.exists(filename):
                new_filename = filename
                print('Warning: overwriting', filename)
            else:
                print(
                    'Warning: filenames similar to', filename, 'already exists in ', arg_dict['output'], ':\n', 
                    existing_filenames,
                    '\nSaving workbook as', new_filename, 'to avoid confusion.'
                )
        wb.save(new_filename)

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

def append(file, sheet, row:list[str]):
    if file not in meta_data:
        createWorkBook(file, sheet)
    else:
        if not sheet in meta_data[file]['sheets']:
            createSheet(file, sheet)
    # by now we have file and sheet
    ws:Worksheet = meta_data[file]['sheets'][sheet]['object']
    current_row = meta_data[file]['sheets'][sheet]['write_row']
    for idx, field in enumerate(row):
        ws.cell(row=current_row, column=(idx+1), value=convert(field))
    meta_data[file]['sheets'][sheet]['write_row'] = current_row + 1

# init:
# if the workbook is not given
# use default workbook name
# if sheet name not given 
# use default sheet name

# if the workbook name is changed:
# if sheet name not given 
# use default sheet name

# note: no empty workbook names

currentWorkBook = WB_DEFAULT
currentSheet = ST_DEFAULT

try:
    # Read all input from stdin
    input_non_empty_flag = False
    
    for row in csv.reader(sys.stdin):
        
        # create workbook only when there is a line to write
        # note we may have currentSheet as None at this point
        # but currentWorkBook can never be None.

        if arg_dict['split']:
            # dynamicall change destination workbook and destination sheet
            if len(row) < 1:
                print('Error: split mode is used, expecting at least 1 fields in every row containing workbook_name:sheet_name, given row is empty.\nInput row from stdin:', row , file=sys.stderr)
                exit(1)

            # assume first 2 fields are workbook name and sheet name
            # this means every row should have at least 2 fields

            if ':' in row[0]:
                row_workBook, row_sheet = [s.strip() for s in row[0].split(':')]
                if row_workBook:
                    # change currentworkbook if placed.
                    currentWorkBook = row_workBook
                if row_sheet:
                    currentSheet = row_sheet
                row = row[1:]
            else:
                print('In split mode, first field in every row should contain workbook_name:sheet_name', row, file=sys.stderr)

        # append row at right place
        append(currentWorkBook, currentSheet, row)

        input_non_empty_flag = True
    
    save_all_workbooks()

    # Input error handling
    if not input_non_empty_flag:
        print('<empty stream>', file=sys.stderr)
        sys.exit(1)
    
except KeyboardInterrupt:
    print('\nOperation cancelled by user', file=sys.stderr)
    sys.exit(1)

